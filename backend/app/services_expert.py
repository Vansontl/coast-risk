from __future__ import annotations
from collections import defaultdict
from typing import Any, Dict, List, Callable
import openpyxl

ProgressCallback = Callable[[int, str], None] | None


def normalize_text(value: Any) -> str:
    import unicodedata
    s = str(value or '').lower().strip()
    s = unicodedata.normalize('NFD', s)
    s = ''.join(ch for ch in s if unicodedata.category(ch) != 'Mn')
    s = s.replace('đ', 'd')
    return s


def parse_likert(value: Any) -> float | None:
    s = normalize_text(value)
    if 'hiem khi' in s or 'rat thap' in s:
        return 1
    if 'doi khi' in s or 'thap' in s:
        return 2
    if 'thinh thoang' in s or 'trung binh' in s:
        return 3
    if 'thuong xuyen' in s or 'cao' in s:
        return 4
    if 'gan nhu chac chan' in s or 'rat cao' in s:
        return 5
    if 'kiem soat tot' in s:
        return 2
    if 'kiem soat duoc' in s:
        return 3
    if 'kho kiem soat' in s:
        return 4
    if 'gan nhu khong kiem soat duoc' in s:
        return 5
    if 'anh huong nhe' in s:
        return 2
    if 'anh huong vua' in s:
        return 3
    if 'anh huong lon' in s:
        return 4
    if 'anh huong nghiem trong' in s:
        return 5
    try:
        n = float(value)
        if 1 <= n <= 5:
            return n
    except Exception:
        return None
    return None


def group_risk_name(name: str) -> str:
    s = normalize_text(name)
    if any(k in s for k in ['mua', 'bao', 'gio', 'song', 'trieu']):
        return 'Điều kiện tự nhiên'
    if any(k in s for k in ['dia chat', 'nen mong', 'khao sat']):
        return 'Địa chất - nền móng'
    if any(k in s for k in ['ky thuat', 'cong nghe', 'bien phap']):
        return 'Kỹ thuật - công nghệ'
    if any(k in s for k in ['vat lieu', 'vat tu', 'thiet bi', 'logistics', 'van chuyen']):
        return 'Vật liệu - logistics'
    if any(k in s for k in ['to chuc', 'quan ly', 'nha thau', 'phoi hop']):
        return 'Tổ chức - quản lý'
    if any(k in s for k in ['an toan', 'moi truong', 'su co']):
        return 'An toàn - môi trường'
    return 'Khác'


def extract_factor(header: str) -> str:
    import re
    match = re.search(r'\[(.*)\]', str(header or ''))
    return (match.group(1).strip() if match else str(header or '').strip())


def average(values: List[float]) -> float:
    return sum(values) / len(values) if values else 0.0


def parse_expert_survey_excel(path: str, progress: ProgressCallback = None) -> Dict[str, Any]:
    if progress: progress(5, 'Đang mở file Excel khảo sát')
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    rows_iter = ws.iter_rows(values_only=True)
    headers_row = next(rows_iter, None)
    if not headers_row:
        return {'sheet_name': sheet_name, 'response_count': 0, 'results': []}

    headers = [str(h or '').strip() for h in headers_row]
    if progress: progress(20, 'Đang nhận diện cột P/I/C')

    factor_map: Dict[str, Dict[str, int | None]] = defaultdict(lambda: {'P': None, 'I': None, 'C': None})
    for idx, header in enumerate(headers):
        normalized = normalize_text(header)
        factor = extract_factor(header)
        if 'danh gia kha nang xay ra' in normalized:
            factor_map[factor]['P'] = idx
        elif 'danh gia muc do anh huong' in normalized:
            factor_map[factor]['I'] = idx
        elif 'danh gia kha nang kiem soat' in normalized:
            factor_map[factor]['C'] = idx

    valid_factors = {factor: cols for factor, cols in factor_map.items() if cols['P'] is not None and cols['I'] is not None and cols['C'] is not None}
    accum = {factor: {'P': [], 'I': [], 'C': []} for factor in valid_factors}

    response_count = 0
    if progress: progress(35, 'Đang đọc dữ liệu khảo sát')
    for row in rows_iter:
        response_count += 1
        row = list(row)
        for factor, cols in valid_factors.items():
            p = parse_likert(row[cols['P']] if cols['P'] is not None and cols['P'] < len(row) else None)
            i = parse_likert(row[cols['I']] if cols['I'] is not None and cols['I'] < len(row) else None)
            c = parse_likert(row[cols['C']] if cols['C'] is not None and cols['C'] < len(row) else None)
            if p is not None: accum[factor]['P'].append(p)
            if i is not None: accum[factor]['I'].append(i)
            if c is not None: accum[factor]['C'].append(c)
        if progress and response_count % 25 == 0:
            percent = min(75, 35 + response_count // 8)
            progress(percent, f'Đã đọc {response_count} phiếu khảo sát')

    if progress: progress(85, 'Đang tính chỉ số P/I/C/R')
    results = []
    for factor, values in accum.items():
        p = average(values['P'])
        i = average(values['I'])
        c = average(values['C'])
        r = (p + i + c) / 3 if any([p, i, c]) else 0
        results.append({
            'factor': factor,
            'group': group_risk_name(factor),
            'P': round(p, 4),
            'I': round(i, 4),
            'C': round(c, 4),
            'R': round(r, 4),
        })

    results.sort(key=lambda x: x['R'], reverse=True)
    for index, item in enumerate(results, start=1):
        item['rank'] = index

    if progress: progress(100, 'Hoàn tất phân tích khảo sát chuyên gia')
    return {
        'sheet_name': sheet_name,
        'response_count': response_count,
        'results': results,
    }
