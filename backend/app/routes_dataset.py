from __future__ import annotations
import json
from pathlib import Path
from uuid import uuid4
from fastapi import APIRouter, UploadFile, File, HTTPException, Depends, BackgroundTasks, Form
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from .config import settings
from .database import get_db
from .models import Dataset, DatasetRow, Region
from .progress_store import create_job, update_job
from .worker_jobs import process_dataset_job
from .schemas import DatasetRowUpsert
from .services_risk_labeling import auto_label_risk
import openpyxl

router = APIRouter(prefix='/api/datasets', tags=['datasets'])


FULL_DATASET_COLUMNS = [
    'projectId', 'region', 'projectName', 'waveHeight', 'tideMode', 'windSpeed', 'stormLevel',
    'soilType', 'weakLayer', 'slideRisk', 'surveyQuality', 'techComplex', 'constructionDiff',
    'equipmentDepend', 'techError', 'materialSupply', 'equipmentMobilize', 'transportRisk',
    'resourceShortage', 'siteManage', 'coordinationRisk', 'scheduleRisk', 'issueHandling',
    'laborSafety', 'marineSafety', 'environmentRisk', 'emergencyResponse', 'riskScore', 'riskLevel'
]

DATASET_TEMPLATE_HEADERS = [
    'Mã công trình', 'Vùng', 'Tên công trình', 'Chiều cao sóng', 'Chế độ thủy triều', 'Tốc độ gió', 'Mức độ ảnh hưởng bão',
    'Loại đất nền', 'Chiều dày lớp đất yếu', 'Nguy cơ trượt mái', 'Chất lượng khảo sát', 'Độ phức tạp kỹ thuật', 'Độ khó biện pháp thi công',
    'Phụ thuộc thiết bị chuyên dụng', 'Nguy cơ sai sót kỹ thuật', 'Khả năng cung ứng vật liệu', 'Khả năng huy động thiết bị', 'Rủi ro vận chuyển',
    'Nguy cơ thiếu hụt nguồn lực', 'Năng lực quản lý hiện trường', 'Rủi ro phối hợp các bên', 'Nguy cơ chậm tiến độ', 'Khả năng xử lý sự cố',
    'An toàn lao động', 'An toàn thi công trên biển', 'Rủi ro môi trường', 'Khả năng ứng cứu khẩn cấp', 'Điểm rủi ro', 'Mức rủi ro'
]


def serialize_row(row: DatasetRow):
    if row.payload_json:
        data = json.loads(row.payload_json)
    else:
        data = {
            'projectId': row.project_id,
            'projectName': row.project_name,
            'region': row.region_name,
            'waveHeight': row.wave_height,
            'tideMode': row.tide_mode,
            'windSpeed': row.wind_speed,
            'stormLevel': row.storm_level,
            'soilType': row.soil_type,
            'weakLayer': row.weak_layer,
            'slideRisk': row.slide_risk,
            'surveyQuality': row.survey_quality,
            'techComplex': row.tech_complex,
            'constructionDiff': row.construction_diff,
            'equipmentDepend': row.equipment_depend,
            'techError': row.tech_error,
            'materialSupply': row.material_supply,
            'equipmentMobilize': row.equipment_mobilize,
            'transportRisk': row.transport_risk,
            'resourceShortage': row.resource_shortage,
            'siteManage': row.site_manage,
            'coordinationRisk': row.coordination_risk,
            'scheduleRisk': row.schedule_risk,
            'issueHandling': row.issue_handling,
            'laborSafety': row.labor_safety,
            'marineSafety': row.marine_safety,
            'environmentRisk': row.environment_risk,
            'emergencyResponse': row.emergency_response,
            'riskScore': row.risk_score,
            'riskLevel': row.risk_level,
        }
    return {
        'id': row.id,
        'projectId': row.project_id,
        'projectName': row.project_name,
        'riskScore': row.risk_score,
        'riskLevel': row.risk_level,
        'data': data,
    }


@router.post('/upload')
async def upload_dataset(background_tasks: BackgroundTasks, dataset_name: str = Form(''), file: UploadFile = File(...)):
    job_id = str(uuid4())
    create_job(job_id, 'dataset-upload')
    suffix = Path(file.filename or 'dataset.xlsx').suffix or '.xlsx'
    upload_dir = Path(settings.upload_dir).resolve()
    upload_dir.mkdir(parents=True, exist_ok=True)
    save_path = upload_dir / f"dataset-{uuid4()}{suffix}"
    save_path.write_bytes(await file.read())
    update_job(job_id, status='running', percent=5, message='Đã nhận file dataset, chuẩn bị xử lý nền')
    final_dataset_name = (dataset_name or '').strip() or (file.filename or 'uploaded-dataset')
    background_tasks.add_task(process_dataset_job, job_id, final_dataset_name, str(save_path))
    return {
        'jobId': job_id,
        'status': 'queued',
        'message': 'File dataset đã nhận, backend đang xử lý nền'
    }


@router.get('')
def list_datasets(db: Session = Depends(get_db)):
    datasets = db.query(Dataset).order_by(Dataset.id.desc()).all()
    result = []
    for d in datasets:
      first_row = db.query(DatasetRow).filter(DatasetRow.dataset_id == d.id).first()
      row_count = db.query(DatasetRow).filter(DatasetRow.dataset_id == d.id).count()
      region = db.query(Region).filter(Region.id == d.region_id).first()
      result.append({
          'id': d.id,
          'region_id': d.region_id,
          'name': d.name,
          'is_active': d.is_active,
          'row_count': row_count,
          'region_name': first_row.region_name if first_row else (region.name if region else None),
      })
    return result


@router.get('/{dataset_id}/rows')
def list_dataset_rows(dataset_id: int, db: Session = Depends(get_db)):
    rows = db.query(DatasetRow).filter(DatasetRow.dataset_id == dataset_id).all()
    return [serialize_row(row) for row in rows]


@router.post('/{dataset_id}/rows')
def create_dataset_row(dataset_id: int, payload: DatasetRowUpsert, db: Session = Depends(get_db)):
    dataset = db.query(Dataset).filter(Dataset.id == dataset_id).first()
    if not dataset:
        raise HTTPException(status_code=404, detail='Dataset not found')
    labeled = auto_label_risk(payload.data)
    row = DatasetRow(
        dataset_id=dataset_id,
        project_id=payload.projectId,
        project_name=payload.projectName,
        payload_json=json.dumps(labeled, ensure_ascii=False),
        risk_score=labeled.get('riskScore'),
        risk_level=labeled.get('riskLevel'),
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return serialize_row(row)


@router.put('/rows/{row_id}')
def update_dataset_row(row_id: int, payload: DatasetRowUpsert, db: Session = Depends(get_db)):
    row = db.query(DatasetRow).filter(DatasetRow.id == row_id).first()
    if not row:
        raise HTTPException(status_code=404, detail='Dataset row not found')
    labeled = auto_label_risk(payload.data)
    row.project_id = payload.projectId
    row.project_name = payload.projectName
    row.payload_json = json.dumps(labeled, ensure_ascii=False)
    row.risk_score = labeled.get('riskScore')
    row.risk_level = labeled.get('riskLevel')
    db.commit()
    db.refresh(row)
    return serialize_row(row)


@router.get('/template/export')
def export_dataset_template(dataset_name: str = 'dataset_mau'):
    template_source = Path(settings.storage_dir).resolve() / 'templates' / 'du_lieu_mau.xlsx'
    if not template_source.exists():
        raise HTTPException(status_code=404, detail='Không tìm thấy file mẫu dữ liệu')

    output_dir = Path(settings.report_dir).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    safe_name = ''.join(ch if ch.isalnum() or ch in ('_', '-', ' ') else '_' for ch in dataset_name).strip() or 'dataset_mau'
    output_path = output_dir / f"{safe_name}.xlsx"

    wb = openpyxl.load_workbook(template_source)
    wb.save(output_path)

    return FileResponse(
        str(output_path),
        filename=f"{safe_name}.xlsx",
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@router.get('/schema')
def dataset_schema():
    return {
        'columnCount': len(FULL_DATASET_COLUMNS),
        'columns': FULL_DATASET_COLUMNS,
        'templateHeaders': DATASET_TEMPLATE_HEADERS,
    }


@router.post('/import-single-row')
async def import_single_row_from_excel(file: UploadFile = File(...)):
    suffix = Path(file.filename or 'import.xlsx').suffix or '.xlsx'
    upload_dir = Path(settings.upload_dir).resolve()
    upload_dir.mkdir(parents=True, exist_ok=True)
    save_path = upload_dir / f"single-row-{uuid4()}{suffix}"
    save_path.write_bytes(await file.read())

    wb = openpyxl.load_workbook(save_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail='File Excel không có dòng dữ liệu để import')

    headers = [str(v).strip() if v is not None else '' for v in rows[0]]
    values = rows[1]
    mapping = dict(zip(DATASET_TEMPLATE_HEADERS, FULL_DATASET_COLUMNS))
    payload = {}
    for idx, header in enumerate(headers):
        field = mapping.get(header)
        if field:
            payload[field] = values[idx] if idx < len(values) else None

    return {'ok': True, 'row': payload}


@router.post('/delete-bulk')
def delete_bulk_datasets(payload: dict, db: Session = Depends(get_db)):
    ids = payload.get('datasetIds') or []
    if not ids:
        raise HTTPException(status_code=400, detail='No dataset ids provided')
    deleted = []
    for dataset_id in ids:
        dataset = db.query(Dataset).filter(Dataset.id == int(dataset_id)).first()
        if dataset:
            db.query(DatasetRow).filter(DatasetRow.dataset_id == int(dataset_id)).delete()
            db.delete(dataset)
            deleted.append(int(dataset_id))
    db.commit()
    return {'ok': True, 'deletedDatasetIds': deleted}


@router.delete('/{dataset_id}')
def delete_dataset(dataset_id: int, db: Session = Depends(get_db)):
    dataset = db.query(Dataset).filter(Dataset.id == dataset_id).first()
    if not dataset:
        raise HTTPException(status_code=404, detail='Dataset not found')
    db.query(DatasetRow).filter(DatasetRow.dataset_id == dataset_id).delete()
    db.delete(dataset)
    db.commit()
    return {'ok': True, 'deletedDatasetId': dataset_id}


@router.delete('/rows/{row_id}')
def delete_dataset_row(row_id: int, db: Session = Depends(get_db)):
    row = db.query(DatasetRow).filter(DatasetRow.id == row_id).first()
    if not row:
        raise HTTPException(status_code=404, detail='Dataset row not found')
    db.delete(row)
    db.commit()
    return {'ok': True, 'deletedRowId': row_id}
