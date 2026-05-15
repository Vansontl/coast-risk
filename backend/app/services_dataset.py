from __future__ import annotations
from typing import Any, Dict, List, Callable
import openpyxl
from .services_risk_labeling import auto_label_risk

ProgressCallback = Callable[[int, str], None] | None


def normalize_dataset_row(row: Dict[str, Any]) -> Dict[str, Any]:
    normalized = {
        'projectId': row.get('Mã công trình') or row.get('projectId'),
        'region': row.get('Vùng') or row.get('region'),
        'projectName': row.get('Tên công trình') or row.get('projectName'),
        'waveHeight': row.get('Chiều cao sóng') or row.get('waveHeight'),
        'tideMode': row.get('Chế độ thủy triều') or row.get('tideMode'),
        'windSpeed': row.get('Tốc độ gió') or row.get('windSpeed'),
        'stormLevel': row.get('Mức độ ảnh hưởng bão') or row.get('stormLevel'),
        'soilType': row.get('Loại đất nền') or row.get('soilType'),
        'weakLayer': row.get('Chiều dày lớp đất yếu') or row.get('weakLayer'),
        'slideRisk': row.get('Nguy cơ trượt mái') or row.get('slideRisk'),
        'surveyQuality': row.get('Chất lượng khảo sát') or row.get('surveyQuality'),
        'techComplex': row.get('Độ phức tạp kỹ thuật') or row.get('techComplex'),
        'constructionDiff': row.get('Độ khó biện pháp thi công') or row.get('constructionDiff'),
        'equipmentDepend': row.get('Phụ thuộc thiết bị chuyên dụng') or row.get('equipmentDepend'),
        'techError': row.get('Nguy cơ sai sót kỹ thuật') or row.get('techError'),
        'materialSupply': row.get('Khả năng cung ứng vật liệu') or row.get('materialSupply'),
        'equipmentMobilize': row.get('Khả năng huy động thiết bị') or row.get('equipmentMobilize'),
        'transportRisk': row.get('Rủi ro vận chuyển') or row.get('transportRisk'),
        'resourceShortage': row.get('Nguy cơ thiếu hụt nguồn lực') or row.get('resourceShortage'),
        'siteManage': row.get('Năng lực quản lý hiện trường') or row.get('siteManage'),
        'coordinationRisk': row.get('Rủi ro phối hợp các bên') or row.get('coordinationRisk'),
        'scheduleRisk': row.get('Nguy cơ chậm tiến độ') or row.get('scheduleRisk'),
        'issueHandling': row.get('Khả năng xử lý sự cố') or row.get('issueHandling'),
        'laborSafety': row.get('An toàn lao động') or row.get('laborSafety'),
        'marineSafety': row.get('An toàn thi công trên biển') or row.get('marineSafety'),
        'environmentRisk': row.get('Rủi ro môi trường') or row.get('environmentRisk'),
        'emergencyResponse': row.get('Khả năng ứng cứu khẩn cấp') or row.get('emergencyResponse'),
        'riskScore': row.get('Điểm rủi ro') or row.get('riskScore'),
        'riskLevel': row.get('Mức rủi ro') or row.get('riskLevel'),
    }
    return auto_label_risk(normalized)


def required_columns() -> List[str]:
    return [
        'projectId', 'region', 'projectName', 'waveHeight', 'tideMode', 'windSpeed', 'stormLevel',
        'soilType', 'weakLayer', 'slideRisk', 'surveyQuality', 'techComplex', 'constructionDiff',
        'equipmentDepend', 'techError', 'materialSupply', 'equipmentMobilize', 'transportRisk',
        'resourceShortage', 'siteManage', 'coordinationRisk', 'scheduleRisk', 'issueHandling',
        'laborSafety', 'marineSafety', 'environmentRisk', 'emergencyResponse', 'riskScore', 'riskLevel'
    ]


def validate_rows(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    if not rows:
        return {'ok': False, 'missing': required_columns(), 'count': 0, 'rows': []}
    normalized = [normalize_dataset_row(row) for row in rows]
    first = normalized[0]
    missing = [col for col in required_columns() if first.get(col) is None]
    return {'ok': len(missing) == 0, 'missing': missing, 'count': len(first.keys()), 'rows': normalized}


def parse_dataset_excel(path: str, progress: ProgressCallback = None) -> Dict[str, Any]:
    if progress: progress(5, 'Đang mở file Excel dataset')
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)
    headers_row = next(rows_iter, None)
    if not headers_row:
        return {'sheet_name': sheet_name, 'validation': validate_rows([])}
    headers = [str(h or '').strip() for h in headers_row]
    records = []
    if progress: progress(25, 'Đang đọc các dòng dataset')
    count = 0
    for row in rows_iter:
        count += 1
        records.append({headers[i]: row[i] if i < len(row) else None for i in range(len(headers))})
        if progress and count % 50 == 0:
            percent = min(70, 25 + count // 12)
            progress(percent, f'Đã đọc {count} dòng dữ liệu')
    if progress: progress(80, 'Đang chuẩn hóa và kiểm tra cấu trúc dataset')
    validation = validate_rows(records)
    if progress: progress(100, 'Hoàn tất xử lý dataset')
    return {
        'sheet_name': sheet_name,
        'validation': validation,
    }
